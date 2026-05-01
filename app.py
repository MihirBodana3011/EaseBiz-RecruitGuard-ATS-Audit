import os
import pandas as pd
import re
import io
import json
import base64
import threading
import uuid
import time
from flask import Flask, request, jsonify, send_file, Response, stream_with_context
from flask_cors import CORS
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
CORS(app, expose_headers=['X-Audit-Stats'])

# In-memory job store: job_id -> {status, progress, step, message, result_bytes, stats, filename}
jobs = {}

# â”€â”€ Excel row color fills â”€â”€
FILL_GREEN  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Clean
FILL_RED    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Error
FILL_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Warning
FILL_BLUE   = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # Duplicate
FILL_HEADER = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')  # Header
FONT_HEADER = Font(color='FFFFFF', bold=True, size=11)

def get_row_fill(all_errors_val):
    v = str(all_errors_val)
    if v == 'Clean Record': return FILL_GREEN
    if 'Duplicate' in v:   return FILL_BLUE
    return FILL_RED

def apply_sheet_formatting(ws):
    """Bold header, freeze pane, auto column width (header-based, fast)."""
    for cell in ws[1]:
        cell.fill      = FILL_HEADER
        cell.font      = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 25
    # Width based on header text only — fast, no full-column scan
    for cell in ws[1]:
        col_letter = get_column_letter(cell.column)
        hdr_len = len(str(cell.value)) if cell.value else 8
        ws.column_dimensions[col_letter].width = min(hdr_len + 6, 50)

def apply_conditional_coloring(ws, n_rows, last_col_letter):
    """Apply Excel conditional formatting rules — near-instant, no cell iteration."""
    from openpyxl.formatting.rule import FormulaRule
    data_range = f'A2:{last_col_letter}{n_rows + 1}'
    lc = f'${last_col_letter}'
    # Order matters: first matching rule wins in Excel (blue > green > red)
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("Duplicate",{lc}2))'], fill=FILL_BLUE
    ))
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'{lc}2="Clean Record"'], fill=FILL_GREEN
    ))
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'AND({lc}2<>"",{lc}2<>"Clean Record")'], fill=FILL_RED
    ))

# Pre-compiled regexes for maximum performance
RE_MULTIPLE = re.compile(r'[,/|;]|\band\b', re.IGNORECASE)
RE_NON_NUMERIC = re.compile(r'\D')
RE_MOBILE_PATTERN = re.compile(r'(?:(?:\+|00)?(?:91|091|0)[\s\-\(\)\.]*)?([6789](?:[\s\-\(\)\.]*\d){9})')
RE_DUMMY_MOBILE_SEQ = re.compile(r'(1234567|2345678|3456789|4567890|0987654|9876543|8765432|7654321)')
RE_REPEATED_DIGITS = re.compile(r'(.)\1{6,}')
RE_REPEATING_ZEROS = re.compile(r'0{5,}$')
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

RE_EMAIL_FORMAT = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
RE_NAME_SPAM = re.compile(r'(.)\1{4,}')
RE_NAME_STARTS_WITH_NUM = re.compile(r'^\d')

# Comprehensive designation / role / profile / industry keywords to detect in name fields
DESIGNATION_KEYWORDS = [

    # â”€â”€ SENIORITY / LEVEL PREFIXES & SUFFIXES â”€â”€
    'senior', 'junior', 'associate', 'assistant', 'principal', 'staff', 'lead',
    'sr.', 'jr.', 'sr ', 'jr ', 'l1', 'l2', 'l3', 'l4', 'l5', 'l6',
    'entry level', 'mid level', 'mid-level', 'senior level', 'executive level',
    'head of', 'chief', 'vp', 'avp', 'svp', 'evp', 'gm', 'dgm', 'agm',

    # â”€â”€ C-SUITE / TOP LEADERSHIP â”€â”€
    'ceo', 'cto', 'cfo', 'coo', 'ciso', 'cmo', 'chro', 'cpo', 'cdo', 'cro',
    'chief executive', 'chief technology', 'chief financial', 'chief operating',
    'chief marketing', 'chief human', 'chief product', 'chief data', 'chief revenue',
    'managing director', 'executive director', 'non executive', 'chairman', 'chairperson',
    'founder', 'co-founder', 'cofounder', 'owner', 'proprietor', 'partner',
    'president', 'vice president', 'director', 'general manager',

    # â”€â”€ MANAGEMENT â”€â”€
    'manager', 'management', 'supervisor', 'coordinator', 'administrator',
    'team lead', 'team leader', 'group lead', 'module lead', 'tech lead',
    'project manager', 'program manager', 'product manager', 'delivery manager',
    'account manager', 'branch manager', 'area manager', 'zonal manager',
    'regional manager', 'cluster manager', 'territory manager', 'city manager',
    'operations manager', 'service manager', 'relationship manager',
    'assistant manager', 'deputy manager', 'senior manager',

    # â”€â”€ IT / SOFTWARE / TECH â”€â”€
    'engineer', 'developer', 'programmer', 'coder', 'architect', 'analyst',
    'consultant', 'technician', 'specialist', 'administrator', 'support',
    'software engineer', 'software developer', 'software architect',
    'frontend', 'back end', 'backend', 'front end', 'full stack', 'fullstack',
    'web developer', 'web designer', 'mobile developer', 'app developer',
    'android developer', 'ios developer', 'react developer', 'angular developer',
    'node developer', 'python developer', 'java developer', 'php developer',
    'dot net', 'dotnet', '.net developer', 'golang', 'ruby', 'scala developer',
    'devops', 'devsecops', 'mlops', 'dataops', 'sre', 'site reliability',
    'cloud engineer', 'cloud architect', 'aws', 'azure', 'gcp', 'google cloud',
    'infrastructure', 'network engineer', 'network admin', 'system admin',
    'sysadmin', 'it admin', 'it support', 'it executive', 'it officer',
    'hardware engineer', 'embedded', 'firmware', 'iot engineer',
    'cybersecurity', 'security engineer', 'infosec', 'penetration tester',
    'ethical hacker', 'vapt', 'soc analyst', 'security analyst',
    'data scientist', 'data analyst', 'data engineer', 'data architect',
    'business analyst', 'systems analyst', 'bi analyst', 'bi developer',
    'etl developer', 'database admin', 'dba', 'sql developer', 'oracle dba',
    'machine learning', 'deep learning', 'ai engineer', 'ml engineer',
    'nlp engineer', 'computer vision', 'data modeler',
    'qa engineer', 'qa analyst', 'qa tester', 'quality assurance',
    'quality engineer', 'test engineer', 'automation tester', 'manual tester',
    'performance tester', 'selenium', 'appium', 'cypress',
    'scrum master', 'agile coach', 'product owner', 'release manager',
    'build engineer', 'configuration manager', 'erp consultant', 'sap consultant',
    'salesforce', 'servicenow', 'workday consultant', 'oracle consultant',
    'technical writer', 'documentation', 'it recruiter', 'technical recruiter',

    # â”€â”€ DESIGN / CREATIVE / MEDIA â”€â”€
    'designer', 'graphic designer', 'ui designer', 'ux designer', 'ui/ux',
    'product designer', 'visual designer', 'interaction designer',
    'motion designer', 'brand designer', 'logo designer', 'illustrator',
    'animator', '3d artist', '2d artist', 'vfx artist', 'video editor',
    'content creator', 'content writer', 'copywriter', 'copy editor',
    'scriptwriter', 'creative director', 'art director', 'creative head',
    'photographer', 'videographer', 'cinematographer', 'production designer',
    'social media', 'social media manager', 'digital marketer', 'influencer',
    'blogger', 'podcaster', 'youtuber', 'editor', 'sub editor', 'journalist',
    'reporter', 'news anchor', 'anchor', 'rj', 'vj', 'media', 'broadcast',

    # â”€â”€ MARKETING / SALES / BUSINESS DEVELOPMENT â”€â”€
    'marketing', 'marketer', 'digital marketing', 'performance marketing',
    'growth hacker', 'seo', 'sem', 'ppc', 'google ads', 'email marketing',
    'affiliate marketing', 'brand manager', 'brand executive',
    'sales', 'sales executive', 'sales officer', 'sales representative',
    'sales engineer', 'inside sales', 'field sales', 'direct sales',
    'telesales', 'telecaller', 'tele caller', 'cold calling',
    'business development', 'bd manager', 'bd executive', 'bdm', 'bde',
    'key account', 'national sales', 'channel sales', 'retail sales',
    'pre sales', 'presales', 'solution consultant', 'bid manager',
    'crm', 'customer success', 'customer experience', 'cx',
    'client servicing', 'client relationship', 'relationship executive',

    # â”€â”€ HUMAN RESOURCES / TALENT â”€â”€
    'hr', 'human resource', 'human resources', 'hrbp', 'hr business partner',
    'hr manager', 'hr executive', 'hr generalist', 'hr specialist',
    'talent acquisition', 'talent management', 'recruiter', 'recruitment',
    'hiring', 'sourcing', 'headhunter', 'staffing', 'workforce planning',
    'learning and development', 'l&d', 'training', 'organization development',
    'od', 'compensation', 'benefits', 'payroll', 'employee relations',
    'hr operations', 'people operations', 'people partner',

    # â”€â”€ FINANCE / ACCOUNTS / BANKING â”€â”€
    'accountant', 'accounts', 'finance', 'financial', 'auditor', 'audit',
    'chartered accountant', 'ca', 'cpa', 'cfa', 'acca', 'icwa', 'cma',
    'tax', 'taxation', 'gst', 'tds', 'direct tax', 'indirect tax',
    'payroll', 'bookkeeper', 'billing', 'invoice', 'accounts payable',
    'accounts receivable', 'credit analyst', 'credit manager',
    'financial analyst', 'financial controller', 'controller',
    'treasury', 'fund manager', 'portfolio manager', 'investment analyst',
    'investment banker', 'banker', 'bank manager', 'relationship manager',
    'loan officer', 'mortgage', 'insurance', 'actuary', 'underwriter',
    'risk analyst', 'risk manager', 'compliance', 'regulatory', 'kyc',
    'aml', 'anti money laundering', 'internal audit', 'statutory audit',

    # â”€â”€ OPERATIONS / SUPPLY CHAIN / LOGISTICS â”€â”€
    'operations', 'supply chain', 'logistics', 'procurement', 'purchasing',
    'buyer', 'vendor management', 'sourcing manager', 'inventory',
    'warehouse', 'warehouse manager', 'warehouse executive', 'storekeeper',
    'distribution', 'dispatch', 'delivery', 'transport', 'fleet',
    'import', 'export', 'customs', 'freight', 'shipping', 'packaging',
    'production', 'production manager', 'plant manager', 'plant head',
    'factory manager', 'quality control', 'quality manager', 'qc',
    'quality assurance', 'process engineer', 'industrial engineer',
    'lean', 'six sigma', 'kaizen', 'maintenance engineer', 'maintenance manager',

    # â”€â”€ HEALTHCARE / MEDICAL / PHARMA â”€â”€
    'doctor', 'dr.', 'physician', 'surgeon', 'specialist', 'consultant doctor',
    'dentist', 'dermatologist', 'cardiologist', 'neurologist', 'oncologist',
    'radiologist', 'pathologist', 'anesthesiologist', 'orthopedic',
    'gynecologist', 'pediatrician', 'psychiatrist', 'ophthalmologist',
    'nurse', 'registered nurse', 'staff nurse', 'nursing',
    'pharmacist', 'pharmacy', 'medical representative', 'mr',
    'medical officer', 'clinical', 'clinical research', 'clinical trial',
    'lab technician', 'lab analyst', 'phlebotomist', 'radiographer',
    'physiotherapist', 'therapist', 'occupational therapist', 'dietitian',
    'nutritionist', 'paramedic', 'emt', 'healthcare', 'medical',
    'hospital administrator', 'ward manager', 'ward sister',
    'pharma', 'pharmaceutical', 'drug', 'biotech', 'bioinformatics',

    # â”€â”€ EDUCATION / ACADEMIA â”€â”€
    'teacher', 'professor', 'lecturer', 'faculty', 'tutor', 'instructor',
    'academic', 'researcher', 'research scholar', 'phd', 'post doc',
    'principal', 'vice principal', 'dean', 'provost', 'rector',
    'school teacher', 'preschool teacher', 'kindergarten teacher',
    'curriculum developer', 'education consultant', 'edtech',
    'teaching assistant', 'lab assistant', 'demonstrator',
    'coach', 'mentor', 'counselor', 'career counselor', 'guidance counselor',

    # â”€â”€ LEGAL â”€â”€
    'lawyer', 'advocate', 'attorney', 'solicitor', 'barrister',
    'legal', 'legal counsel', 'legal advisor', 'corporate counsel',
    'paralegal', 'legal executive', 'legal assistant',
    'company secretary', 'cs', 'compliance officer', 'legal manager',
    'judge', 'magistrate', 'notary', 'arbitrator', 'mediator',
    'patent agent', 'intellectual property', 'ip lawyer',

    # â”€â”€ REAL ESTATE / CONSTRUCTION / CIVIL â”€â”€
    'civil engineer', 'structural engineer', 'site engineer', 'site supervisor',
    'project engineer', 'construction manager', 'project coordinator',
    'quantity surveyor', 'qs', 'estimator', 'billing engineer',
    'architect', 'interior designer', 'interior decorator',
    'real estate', 'property consultant', 'property manager',
    'facility manager', 'facility management', 'fm',
    'surveyor', 'town planner', 'urban planner', 'bim', 'autocad',
    'mep engineer', 'mechanical engineer', 'electrical engineer',

    # â”€â”€ MANUFACTURING / ENGINEERING DOMAINS â”€â”€
    'mechanical', 'electrical', 'electronics', 'instrumentation',
    'chemical engineer', 'production engineer', 'manufacturing engineer',
    'design engineer', 'r&d engineer', 'research engineer',
    'tool engineer', 'die designer', 'cnc', 'cad', 'cam',
    'welding engineer', 'quality engineer', 'reliability engineer',
    'ehs', 'safety officer', 'safety engineer', 'environment',
    'energy manager', 'utilities engineer', 'hvac', 'plumber', 'electrician',
    'fitter', 'machinist', 'welder', 'operator', 'supervisor',

    # â”€â”€ HOSPITALITY / TRAVEL / TOURISM â”€â”€
    'chef', 'head chef', 'sous chef', 'executive chef', 'pastry chef',
    'cook', 'baker', 'food', 'food technologist', 'food safety',
    'hotel manager', 'front office', 'front desk', 'receptionist',
    'concierge', 'housekeeping', 'housekeeping supervisor',
    'f&b', 'food and beverage', 'banquet', 'catering',
    'travel agent', 'travel consultant', 'tour guide', 'tour manager',
    'reservation', 'ticketing', 'cabin crew', 'flight attendant',
    'ground staff', 'airline', 'aviation', 'pilot', 'co-pilot',
    'spa', 'wellness', 'gym trainer', 'fitness trainer',

    # â”€â”€ RETAIL / ECOMMERCE / CUSTOMER SERVICE â”€â”€
    'retail', 'store manager', 'floor manager', 'showroom manager',
    'merchandiser', 'visual merchandiser', 'category manager',
    'buyer', 'planner', 'ecommerce', 'marketplace', 'amazon seller',
    'customer service', 'customer support', 'customer care',
    'call center', 'bpo', 'kpo', 'lpo', 'process associate',
    'team leader', 'quality analyst', 'escalation',
    'help desk', 'service desk', 'technical support', 'l1 support', 'l2 support',

    # â”€â”€ GOVERNMENT / PUBLIC SECTOR / DEFENSE â”€â”€
    'ias', 'ips', 'ifs', 'irs', 'government officer', 'civil servant',
    'sub inspector', 'inspector', 'deputy collector', 'collector',
    'tehsildar', 'panchayat', 'municipal', 'clerk', 'typist',
    'police', 'constable', 'army', 'navy', 'air force', 'soldier',
    'defence', 'defense', 'paramilitary', 'crpf', 'bsf', 'cisf',
    'public sector', 'psu', 'sarkari', 'peon', 'ward boy',

    # â”€â”€ MEDIA / ENTERTAINMENT / SPORTS â”€â”€
    'actor', 'actress', 'model', 'performer', 'entertainer',
    'producer', 'director', 'film director', 'casting director',
    'writer', 'screenwriter', 'scriptwriter', 'author', 'novelist',
    'musician', 'singer', 'composer', 'dj', 'sound engineer',
    'dancer', 'choreographer', 'event manager', 'event coordinator',
    'sports coach', 'cricket', 'footballer', 'athlete', 'player',
    'sports analyst', 'fitness', 'yoga instructor',

    # â”€â”€ AGRICULTURE / ENVIRONMENT â”€â”€
    'agronomist', 'agriculture', 'horticulture', 'farm manager',
    'animal husbandry', 'veterinary', 'vet', 'fisheries',
    'environment', 'environmental engineer', 'ecology', 'forest',

    # â”€â”€ GENERAL JOB PROFILE WORDS â”€â”€
    'executive', 'officer', 'associate', 'specialist', 'expert',
    'professional', 'technologist', 'scientist', 'engineer', 'analyst',
    'advisor', 'consultant', 'strategist', 'planner', 'controller',
    'intern', 'fresher', 'trainee', 'apprentice', 'probationer',
    'freelancer', 'contractor', 'part time', 'full time', 'contractual',
    'remote', 'work from home', 'wfh', 'hybrid', 'on site', 'onsite',
    'permanent', 'temporary', 'ad hoc',

    # â”€â”€ EXPERIENCE TAGS PEOPLE PUT IN NAMES â”€â”€
    '0-1 year', '1-2 year', '2-3 year', '3-5 year', '5+ year',
    'experienced', 'exp.', 'years exp', 'yrs exp', 'year experience',
    'looking for', 'seeking', 'open to work', 'available', 'job seeker',
    'job seeking', 'actively looking', 'immediate joiner', 'notice period',

    # â”€â”€ COMMON SHORT ROLE CODES â”€â”€
    'qa', 'ba', 'pm', 'pmo', 'hr', 'bd', 'rm', 'am', 'tm', 'sm',
    'asm', 'rsm', 'nsm', 'csm', 'ssm', 'kam', 'bde', 'bdm', 'sdm',
    'etl', 'dba', 'sap', 'erp', 'crm', 'mrp',
]


# ── Ultra-fast designation lookup: NO regex, pure set/string ops ──
# Single-word keywords → frozenset for O(1) hash lookup
DESIG_WORD_SET = frozenset(k for k in DESIGNATION_KEYWORDS if ' ' not in k)
# Multi-word phrases → tuple, sorted longest-first for early exit
DESIG_PHRASES  = tuple(sorted((k for k in DESIGNATION_KEYWORDS if ' ' in k), key=len, reverse=True))

# Pre-compiled irrelevant terms set
IRRELEVANT_TERMS = frozenset(['test','abc','xyz','na','n/a','unknown','null','none','---','...','.',  'dummy'])


def clean_phone_numbers(mobile_str):
    if pd.isna(mobile_str) or str(mobile_str).strip().lower() == 'nan' or not str(mobile_str).strip():
        return [], "Blank"
        
    p_str = str(mobile_str).strip()
    if p_str.endswith('.0'): p_str = p_str[:-2]
    
    matches = RE_MOBILE_PATTERN.findall(p_str)
    
    valid_numbers = []
    for match in matches:
        num = RE_NON_NUMERIC.sub('', match)
        if num and len(num) == 10:
            valid_numbers.append(num)
            
    seen = set()
    unique_numbers = []
    for num in valid_numbers:
        if num not in seen:
            if RE_DUMMY_MOBILE_SEQ.search(num):
                continue
            if RE_REPEATED_DIGITS.search(num):
                continue
            if RE_REPEATING_ZEROS.search(num):
                continue
            seen.add(num)
            unique_numbers.append(num)
            
    if not unique_numbers:
        return [], "Invalid Dummy/Repeating Number"
        
    return unique_numbers, None

def is_irrelevant_name(name):
    if pd.isna(name) or str(name).strip() == "":
        return "Blank"
    
    name_str = str(name).strip().lower()
    if len(name_str) < 2:
        return "Too Short"
        
    if '@' in name_str:
        return "Contains Email"
    
    if any(char in name_str for char in ['?', '!', '*', '#', '$', '%']):
        return "Contains Garbage Characters"
        
    if name_str.replace(" ", "").isdigit():
        return "Numeric Name"
        
    if RE_NAME_STARTS_WITH_NUM.match(name_str):
        return "Starts with Number"
        
    if RE_NAME_SPAM.search(name_str):
        return "Spam Name"

    if name_str in IRRELEVANT_TERMS:
        return "Irrelevant Term"

    # Ultra-fast designation check: set intersection (O(1)) + phrase scan
    words = set(name_str.split())
    if words & DESIG_WORD_SET:            # single-word match – hash lookup
        return "Contains Designation/Role"
    for phrase in DESIG_PHRASES:          # multi-word match – simple 'in'
        if phrase in name_str:
            return "Contains Designation/Role"

    return None


def validate_email_format(email):
    if pd.isna(email) or str(email).strip() == "":
        return "Blank"
    
    email_str = str(email).strip().lower()
    
    if RE_MULTIPLE.search(email_str):
        return "Multiple Emails Provided"
        
    dummy_emails = ['na@na.com', 'test@test.com', 'abc@xyz.com', 'none@none.com', 'nil@nil.com', 'null@null.com']
    if email_str in dummy_emails:
        return "Dummy/Irrelevant Email"
        
    if not RE_EMAIL_FORMAT.match(email_str):
        return "Invalid Format"
    
    return None

def find_column(df, possible_names):
    for col in df.columns:
        if any(p.lower() in col.lower() for p in possible_names):
            return col
    return None

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    
    try:
        if file.filename.endswith('.csv'):
            try:
                df = pd.read_csv(file, encoding_errors='replace', dtype=str)
            except Exception:
                file.seek(0)
                try:
                    df = pd.read_csv(file, encoding='latin-1', encoding_errors='replace', dtype=str)
                except Exception:
                    file.seek(0)
                    df = pd.read_csv(file, encoding='cp1252', encoding_errors='replace', dtype=str)
        else:
            df = pd.read_excel(file, dtype=str)
        
        preview_data = df.head(20).fillna("").to_dict(orient='records')
        return jsonify({
            "filename": file.filename,
            "rowCount": len(df),
            "preview": preview_data,
            "columns": list(df.columns)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def run_audit(file_bytes, filename, column_mapping, validation_config, job_id):
    """Background processing with chunked progress updates."""
    def upd(step, pct, msg):
        if job_id in jobs:
            jobs[job_id].update({'step': step, 'progress': pct, 'message': msg})
    try:
        upd('parsing', 5, 'Reading file...')
        buf = io.BytesIO(file_bytes)
        if filename.endswith('.csv'):
            try:    df = pd.read_csv(buf, dtype=str, encoding_errors='replace')
            except Exception:
                buf.seek(0)
                try:    df = pd.read_csv(buf, encoding='latin-1', dtype=str, encoding_errors='replace')
                except Exception:
                    buf.seek(0)
                    df = pd.read_csv(buf, encoding='cp1252', dtype=str, encoding_errors='replace')
        else:
            df = pd.read_excel(buf, dtype=str)
        df = df.fillna('')

        # --- Column mapping (user-defined or auto-detect) ---
        cm = column_mapping or {}
        def col_pick(key, fallbacks):
            v = cm.get(key, '')
            if v and v in df.columns: return v
            return find_column(df, fallbacks)

        name_col        = col_pick('name',        ['name','candidate','full name'])
        mobile_col      = col_pick('mobile',      ['mobile','phone','contact'])
        email_col       = col_pick('email',       ['email','mail'])
        resume_col      = col_pick('resume',      ['resume','attached','cv'])
        gender_col      = col_pick('gender',      ['gender'])
        created_by_col  = col_pick('created_by',  ['createdby','created by'])
        skills_col      = col_pick('skills',      ['key_skills','key skills','skills'])
        designation_col = col_pick('designation', ['designation'])
        experience_col  = col_pick('experience',  ['total_experi','total experi','experience'])
        location_col    = col_pick('location',    ['current_loca','current loca','location','city'])

        # --- validation toggles (default all True) ---
        vc = validation_config or {}
        do_name   = vc.get('check_name',   True)
        do_mobile = vc.get('check_mobile', True)
        do_email  = vc.get('check_email',  True)
        do_resume = vc.get('check_resume', True)
        do_gender = vc.get('check_gender', True)
        do_skills = vc.get('check_skills', True)
        do_desig  = vc.get('check_designation_col', True)
        do_exp    = vc.get('check_experience', True)
        do_loc    = vc.get('check_location', True)
        do_desig_in_name = vc.get('flag_designation_in_name', True)

        records = df.to_dict('records')
        total   = len(records)

        upd('counting', 10, 'Counting duplicates...')
        mobile_counts, email_counts = {}, {}
        for row in records:
            if mobile_col and do_mobile:
                extracted, _ = clean_phone_numbers(str(row.get(mobile_col,'')))
                for m in extracted:
                    mobile_counts[m] = mobile_counts.get(m,0) + 1
            if email_col and do_email:
                ev = str(row.get(email_col,'')).strip().lower()
                if ev: email_counts[ev] = email_counts.get(ev,0) + 1

        output_data, max_mobiles, total_valid, total_errors = [], 0, 0, 0
        CHUNK = 5000  # larger chunks = less SSE overhead

        for chunk_start in range(0, total, CHUNK):
            chunk_end = min(chunk_start + CHUNK, total)
            pct = 15 + int((chunk_end / total) * 65)
            upd('processing', pct, f'Processed {chunk_end:,} of {total:,} records...')

            for row in records[chunk_start:chunk_end]:
                try:
                    row_errors = []
                    if all(str(v).strip() == '' for v in row.values()):
                        rc = row.copy()
                        rc['Name Check']   = 'Blank'
                        rc['Mobile Check'] = 'Blank'
                        rc['Email Check']  = 'Blank'
                        if resume_col and do_resume: rc['Resume Check'] = 'Blank'
                        if gender_col and do_gender: rc['Gender Check'] = 'Blank'
                        if created_by_col:           rc['CreatedBy Check'] = 'Blank'
                        if skills_col and do_skills: rc['Skills Check'] = 'Blank'
                        if designation_col and do_desig: rc['Designation Check'] = 'Blank'
                        if experience_col and do_exp: rc['Experience Check'] = 'Blank'
                        if location_col and do_loc:  rc['Location Check'] = 'Blank'
                        rc['All Errors'] = 'Fully Blank Record'
                        output_data.append(rc)
                        total_errors += 1
                        continue

                    # Name
                    name_val = str(row.get(name_col,'')).strip() if name_col else ''
                    if do_name:
                        if do_desig_in_name:
                            n_issue = is_irrelevant_name(name_val)
                        else:
                            n_issue = is_irrelevant_name_basic(name_val)
                        if n_issue: row_errors.append(f'Name: {n_issue}')
                    else:
                        n_issue = None

                    # Mobile
                    raw_mobile = str(row.get(mobile_col,'')).strip() if mobile_col else ''
                    final_mobiles, m_issue = [], None
                    if do_mobile:
                        extracted_mobiles, m_issue = clean_phone_numbers(raw_mobile)
                        for m in extracted_mobiles:
                            if mobile_counts.get(m,0) > 1:
                                row_errors.append(f'Mobile Duplicate ({m})')
                                m_issue = 'Contains Duplicate'
                            else:
                                final_mobiles.append(m)
                        if m_issue and 'Duplicate' not in str(m_issue):
                            row_errors.append(f'Mobile: {m_issue}')
                    max_mobiles = max(max_mobiles, len(final_mobiles))

                    # Email
                    email_val = str(row.get(email_col,'')).strip() if email_col else ''
                    e_issue = None
                    if do_email:
                        e_issue = validate_email_format(email_val)
                        if not e_issue and email_val:
                            if email_counts.get(email_val.lower(),0) > 1:
                                e_issue = 'Duplicate'
                        if e_issue: row_errors.append(f'Email: {e_issue}')

                    # Resume
                    r_issue = None
                    if resume_col and do_resume:
                        res_val = str(row.get(resume_col,'')).strip().lower()
                        if res_val in ['no','n','','null','none','nan']:
                            r_issue = 'Missing'
                            row_errors.append('Resume Missing')

                    rc = row.copy()
                    if do_name:   rc['Name Check']   = n_issue if n_issue else 'Valid'
                    if do_mobile: rc['Mobile Check'] = m_issue if m_issue else 'Valid'
                    if do_email:  rc['Email Check']  = e_issue if e_issue else 'Valid'
                    if resume_col and do_resume:
                        rc['Resume Check'] = r_issue if r_issue else 'Valid'

                    extra = [
                        ('Gender',      gender_col,      do_gender),
                        ('CreatedBy',   created_by_col,  True),
                        ('Skills',      skills_col,      do_skills),
                        ('Designation', designation_col, do_desig),
                        ('Experience',  experience_col,  do_exp),
                        ('Location',    location_col,    do_loc),
                    ]
                    for col_name, col_key, enabled in extra:
                        if col_key and enabled:
                            val = str(row.get(col_key,'')).strip().lower()
                            if val in ['','nan','null','none','n/a']:
                                row_errors.append(f'{col_name} Missing')
                                rc[f'{col_name} Check'] = 'Blank'
                            else:
                                rc[f'{col_name} Check'] = 'Valid'

                    rc['_temp_mobiles'] = final_mobiles
                    rc['All Errors'] = ' | '.join(row_errors) if row_errors else 'Clean Record'
                    if not row_errors: total_valid += 1
                    else:              total_errors += 1
                    output_data.append(rc)
                except Exception as ex:
                    rc = row.copy()
                    rc['All Errors'] = f'Processing Error: {ex}'
                    output_data.append(rc)
                    total_errors += 1

        upd('building', 82, 'Building report columns...')
        for row in output_data:
            if '_temp_mobiles' in row:
                mobiles = row.pop('_temp_mobiles')
                for i in range(max_mobiles):
                    cname = f'Cleaned Mobile {i+1}' if max_mobiles > 1 else 'Cleaned Mobile No'
                    row[cname] = mobiles[i] if i < len(mobiles) else ''

        if output_data:
            df_final = pd.DataFrame(output_data)
            cols = df_final.columns.tolist()
            special_cols = []
            if do_name:   special_cols.append('Name Check')
            if do_mobile: special_cols.append('Mobile Check')
            if do_email:  special_cols.append('Email Check')
            if resume_col and do_resume: special_cols.append('Resume Check')
            for col_name, col_key, enabled in [
                ('Gender',gender_col,do_gender),('CreatedBy',created_by_col,True),
                ('Skills',skills_col,do_skills),('Designation',designation_col,do_desig),
                ('Experience',experience_col,do_exp),('Location',location_col,do_loc)
            ]:
                if col_key and enabled: special_cols.append(f'{col_name} Check')
            for i in range(max_mobiles):
                special_cols.append(f'Cleaned Mobile {i+1}' if max_mobiles > 1 else 'Cleaned Mobile No')
            special_cols.append('All Errors')
            for c in special_cols:
                if c in cols: cols.append(cols.pop(cols.index(c)))
            df_final = df_final[cols]
            def rm_illegal(val):
                if isinstance(val, str): return ILLEGAL_CHARACTERS_RE.sub('', val)
                return val
            df_final = df_final.apply(lambda c: c.map(rm_illegal))
        else:
            df_final, special_cols = pd.DataFrame(), []

        upd('excel', 88, 'Generating Excel report...')
        summary_data = [
            {'Metric':'Total Records Processed','Count':total},
            {'Metric':'Perfectly Clean Records','Count':total_valid},
            {'Metric':'Records with Errors','Count':total_errors},
            {'Metric':'','Count':''}
        ]
        breakdown_dict = {}
        for c in special_cols:
            if c.endswith('Check') and c in df_final.columns:
                mn = c.replace(' Check','')
                summary_data.append({'Metric':f'--- {mn} Breakdown ---','Count':''})
                cnts = df_final[c].value_counts().to_dict()
                breakdown_dict[mn] = cnts
                for k,v in cnts.items():
                    summary_data.append({'Metric':f'  • {k}','Count':v})
                summary_data.append({'Metric':'','Count':''})

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Audit Summary
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Audit Summary', index=False)
            ws_sum = writer.sheets['Audit Summary']
            apply_sheet_formatting(ws_sum)

            # Sheet 2: Audited Data — fast pandas write, then conditional coloring
            df_final.to_excel(writer, sheet_name='Audited Data', index=False)
            ws_data = writer.sheets['Audited Data']
            apply_sheet_formatting(ws_data)
            # Apply color rules (near-instant — Excel renders them on open)
            n_data_rows = len(df_final)
            last_col_letter = get_column_letter(len(df_final.columns))
            apply_conditional_coloring(ws_data, n_data_rows, last_col_letter)


        stats = {'total':total,'errors':total_errors,'valid':total_valid,'breakdown':breakdown_dict}
        jobs[job_id].update({
            'status':'done', 'progress':100, 'step':'done',
            'message':'Audit complete!',
            'result_bytes': output.getvalue(),
            'stats': stats,
        })

    except Exception as ex:
        import traceback; traceback.print_exc()
        if job_id in jobs:
            jobs[job_id].update({'status':'error','error':str(ex)})


def is_irrelevant_name_basic(name):
    """Name check WITHOUT designation/role keyword detection (used when toggle is off)."""
    if pd.isna(name) or str(name).strip() == '': return 'Blank'
    n = str(name).strip().lower()
    if len(n) < 2: return 'Too Short'
    if '@' in n:   return 'Contains Email'
    if any(c in n for c in ['?','!','*','#','$','%']): return 'Contains Garbage Characters'
    if n.replace(' ','').isdigit(): return 'Numeric Name'
    if RE_NAME_STARTS_WITH_NUM.match(n): return 'Starts with Number'
    if RE_NAME_SPAM.search(n): return 'Spam Name'
    if n in ['test','abc','xyz','na','n/a','unknown','null','none','---','...','.',  'dummy']: return 'Irrelevant Term'
    return None


@app.route('/process', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    try:
        column_mapping   = json.loads(request.form.get('column_mapping',   '{}'))
        validation_config= json.loads(request.form.get('validation_config','{}'))
    except Exception:
        column_mapping, validation_config = {}, {}

    job_id     = str(uuid.uuid4())[:10]
    file_bytes = file.read()
    filename   = file.filename
    jobs[job_id] = {
        'status':'running','progress':0,'step':'starting',
        'message':'Starting audit...','result_bytes':None,
        'stats':None,'error':None,'filename':filename
    }
    t = threading.Thread(target=run_audit,
                         args=(file_bytes, filename, column_mapping, validation_config, job_id),
                         daemon=True)
    t.start()
    return jsonify({'job_id': job_id})


@app.route('/stream/<job_id>')
def stream_progress(job_id):
    def generate():
        while True:
            job = jobs.get(job_id)
            if not job:
                yield f"data: {json.dumps({'status':'error','error':'Job not found'})}\n\n"
                break
            payload = {
                'status':   job.get('status'),
                'progress': job.get('progress', 0),
                'step':     job.get('step', ''),
                'message':  job.get('message', ''),
            }
            if job.get('status') == 'done':   payload['stats'] = job.get('stats', {})
            if job.get('status') == 'error':  payload['error'] = job.get('error', 'Unknown error')
            yield f"data: {json.dumps(payload)}\n\n"
            if job.get('status') in ('done','error'): break
            time.sleep(1.0)

    return Response(stream_with_context(generate()),
                    mimetype='text/event-stream',
                    headers={'Cache-Control':'no-cache','X-Accel-Buffering':'no'})


@app.route('/download/<job_id>')
def download_result(job_id):
    job = jobs.get(job_id)
    if not job or not job.get('result_bytes'):
        return jsonify({'error': 'Result not found or still processing'}), 404
    base = job.get('filename','data').rsplit('.',1)[0]
    return send_file(
        io.BytesIO(job['result_bytes']),
        as_attachment=True,
        download_name=f'Audit_Report_{base}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/')
def index(): return send_file('index.html')

@app.route('/style.css')
def serve_css(): return send_file('style.css')

@app.route('/script.js')
def serve_js(): return send_file('script.js')

@app.route('/easebiz_logo.png')
def serve_logo(): return send_file('easebiz_logo.png')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)

